import openpyxl
import os

def set_title(ws):
    """
    adds the title and do not save the workbook
    :param ws: worksheet
    :return: NoneType
    """
    font = openpyxl.styles.Font(size=15,bold=True)
    for char, title in enumerate(["course_url","faculty","department","description","course_code","credits","terms","time_cost","instructors","specific_description","notes"]):
        cell = f"{chr(char+97)}1"
        ws[cell] = title
        ws.column_dimensions[chr(char+97)].width = 20
        ws[cell].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        ws[cell].font = font

def open_wb():
    """
    open workbook called 'McGill_courses.xlsx' in current location,
    if file does not exist, create one and set up the title
    :return: workbook
    """
    try:
        wb = openpyxl.load_workbook(f"{os.getcwd()}/McGill_courses.xlsx")
    except:
        wb = openpyxl.Workbook()
        print("cannot found target workbook, created 'McGill_courses.xlsx'")
        ws = wb.create_sheet("all_courses", 0)
        set_title(ws)
    return wb


def cover_all(courses = []):
    wb = open_wb()
    try:
        del wb["all_courses"]
    except:
        pass
    try:
        ws = wb.create_sheet("all_courses",0)
    except:
        raise AssertionError("fail to open 'all_courses' worksheet!")
    try:
        for c in courses:
            append_list = [c.course_url,c.faculty,c.offer_by,c.description,c.name,c.credit,"& ".join(c.terms),c.time,"& ".join(c.instructors),c.specific_description]
            ws.append(append_list)
            """add c.notes"""
            row = ws.max_row
            for char, content in enumerate(c.info):
                cell = f"{chr(char+107)}{row}"
                ws[cell] = content

        save(wb)
    except(KeyboardInterrupt):
        save(wb)


def append(c):
    wb = open_wb()
    try:
        ws = wb["all_courses"]
    except:
        raise AssertionError("fail to open 'all_courses' worksheet!")
    try:
        append_list = [c.course_url, c.faculty, c.offer_by, c.description, c.name, c.credit, "& ".join(c.terms), c.time,
                       "& ".join(c.instructors), c.specific_description]
        ws.append(append_list)
        row = ws.max_row
        for char, content in enumerate(c.info):
            cell = f"{chr(char + 107)}{row}"
            ws[cell] = content
        save(wb)
    except(KeyboardInterrupt):
        save(wb)

def save(wb):
    try:
        wb.save("McGill_courses.xlsx")
    except(PermissionError):
        print("cannot save file, please close 'McGill_courses.xlsx'!")

if __name__ == "__main__":
    os.popen('mshta vbscript:msgbox("open main.py, not this one!",,"wrong file")(window.close)')
